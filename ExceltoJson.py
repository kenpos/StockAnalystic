import openpyxl
import json
import sys 

# def convert_to_hierarchy(data):
#     hierarchy_data = []
#     for row_data in data:
#         current_level = hierarchy_data
#         for value in row_data:
#             existing_node = next((node for node in current_level if node['value'] == value), None)
#             if existing_node:
#                 current_level = existing_node['children']
#             else:
#                 new_node = {'value': value, 'children': []}
#                 current_level.append(new_node)
#                 current_level = new_node['children']
#     return hierarchy_data

def process_row(sheet, row_index):
    current_row = sheet[row_index]
    tag = current_row[0].value
    info = current_row[3].value
    element_name = current_row[8].value
    depth = current_row[14].value
    
    return {'tag': tag, 'info': info, 'element_name': element_name, 'depth': depth}

def convert_to_hierarchy(sheet):
    hierarchy_data = []
    print(sheet)
    max_row = sheet.max_row

    # Find root rows (Depth 0)
    root_rows = [row for row in sheet.iter_rows(min_row=2, max_row=max_row) if row[14].value == 0]
    
    for root_row in root_rows:
        root_data = process_row(sheet, root_row[0].row)
        root_data['children'] = convert_children(sheet, root_data['depth'], root_data['tag'])
        hierarchy_data.append(root_data)
    
    return hierarchy_data


def convert_children(sheet, parent_depth, parent_tag):
    children = []
    max_row = sheet.max_row

    # Find child rows for the current parent
    child_rows = [row for row in sheet.iter_rows(min_row=2, max_row=max_row) if row[14].value == parent_depth + 1 and row[0].value == parent_tag]

    for child_row in child_rows:
        child_data = process_row(sheet, child_row[0].row)
        child_data['children'] = convert_children(sheet, child_data['depth'], child_data['tag'])
        children.append(child_data)

    return children

def excel_to_json(input_excel_path, output_json_path):
    # Load Excel file
    wb = openpyxl.load_workbook(input_excel_path)
    # sheet = wb.active

    # 全てのシート名を取得
    all_sheet_names = wb.sheetnames

    # 全てのシートにアクセスしてデータを取得する例
    for sheet_name in all_sheet_names:
        if sheet_name not in "目次" and  sheet_name not in"タクソノミ要素リストについて":
            # 階層構造のJSONに変換
            sheet = wb[sheet_name]
            hierarchy_json_data = convert_to_hierarchy(sheet)
            # JSONデータをファイルに書き込む
            output_json_file = f'{sheet_name}_hierarchy.json'
            # output_json_file = 'output_hierarchy.json'
            with open(output_json_file, 'w', encoding='utf-8') as json_file:
                json.dump(hierarchy_json_data, json_file, ensure_ascii=False, indent=2)
        # current_sheet = wb[sheet_name]
        # print(current_sheet)
        # # 2行目のデータを取得
        # labels = [cell.value for cell in current_sheet[2]]       
        # # データの開始行（3行目）を取得
        # start_row = 3
        
        # # データの終了行を取得
        # end_row = current_sheet.max_row
        
        # # 各行にアクセスして必要な範囲のデータを取得
        # data = []
        # for row in current_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=3):
        #     row_data = [cell.value for cell in row]
        #     data.append(row_data)
    
        # # 階層構造のJSONに変換
        # hierarchy_json_data = convert_to_hierarchy(data)
    
        # # JSONデータをファイルに書き込む
        # output_json_file = f'{sheet_name}_hierarchy.json'
        # with open(output_json_file, 'w', encoding='utf-8') as json_file:
        #     json.dump(hierarchy_json_data, json_file, ensure_ascii=False, indent=2)


def main():
    file = sys.argv[1]
    filename = file.split(".")[1]
    jsonname = filename+".json"
    print(filename,jsonname)
    excel_to_json(file, jsonname)

if __name__ == "__main__":
    main()