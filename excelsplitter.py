from openpyxl import load_workbook
import os
import sys
from concurrent.futures import ThreadPoolExecutor

data = [
    [1, "DEI", "jpdei"],
    [2, "企業内容等の開示に関する内閣府令 第二号様式 有価証券届出書", "jpcrp020000"],
    [3, "企業内容等の開示に関する内閣府令 第二号の二様式 有価証券届出書", "jpcrp020200"],
    [4, "企業内容等の開示に関する内閣府令 第二号の三様式 有価証券届出書", "jpcrp020300"],
    [5, "企業内容等の開示に関する内閣府令 第二号の四様式 有価証券届出書", "jpcrp020400"],
    [6, "企業内容等の開示に関する内閣府令 第二号の五様式 有価証券届出書", "jpcrp020500"],
    [7, "企業内容等の開示に関する内閣府令 第二号の六様式 有価証券届出書", "jpcrp020600"],
    [8, "企業内容等の開示に関する内閣府令 第二号の七様式 有価証券届出書", "jpcrp020700"],
    [9, "企業内容等の開示に関する内閣府令 第三号様式 有価証券報告書", "jpcrp030000"],
    [10, "企業内容等の開示に関する内閣府令 第三号の二様式 有価証券報告書", "jpcrp030200"],
    [11, "企業内容等の開示に関する内閣府令 第四号様式 有価証券報告書", "jpcrp040000"],
    [12, "企業内容等の開示に関する内閣府令 第四号の三様式 四半期報告書", "jpcrp040300"],
    [13, "企業内容等の開示に関する内閣府令 第五号様式 半期報告書", "jpcrp050000"],
    [14, "企業内容等の開示に関する内閣府令 第五号の二様式 半期報告書", "jpcrp050200"],
    [15, "企業内容等の開示に関する内閣府令 第五号の三様式 臨時報告書", "jpcrp050300"],
    [16, "企業内容等の開示に関する内閣府令 第七号様式 有価証券届出書", "jpcrp070000"],
    [17, "企業内容等の開示に関する内閣府令 第七号の四様式 有価証券届出書", "jpcrp070400"],
    [18, "企業内容等の開示に関する内閣府令 第八号様式 有価証券報告書", "jpcrp080000"],
    [19, "企業内容等の開示に関する内閣府令 第九号様式 有価証券報告書", "jpcrp090000"],
    [20, "企業内容等の開示に関する内閣府令 第九号の三様式 四半期報告書", "jpcrp090300"],
    [21, "企業内容等の開示に関する内閣府令 第十号様式 半期報告書", "jpcrp100000"],
    [22, "企業内容等の開示に関する内閣府令 第十一号様式 発行登録書", "jpcrp110000"],
    [23, "企業内容等の開示に関する内閣府令 第十一号の二様式 発行登録書", "jpcrp110200"],
    [24, "企業内容等の開示に関する内閣府令 第十一号の二の二様式 発行登録書", "jpcrp110202"],
    [25, "企業内容等の開示に関する内閣府令 第十二号様式 発行登録追補書類", "jpcrp120000"],
    [26, "企業内容等の開示に関する内閣府令 第十二号の二様式 発行登録追補書類", "jpcrp120200"],
    [27, "企業内容等の開示に関する内閣府令 第十七号様式 自己株券買付状況報告書", "jpcrp170000"],
    [28, "財務計算に関する書類その他の情報の適正性を確保するための体制に関する内閣府令 第一号様式 内部統制報告書", "jpctl010000"],
    [29, "大量保有報告書の追加DEI", "jplvh_大量保有報告書_追加DEI"],
    [30, "株券等の大量保有の状況の開示に関する内閣府令 第一号様式 大量保有報告書", "jplvh010000"],
    [31, "株券等の大量保有の状況の開示に関する内閣府令 第一号及び第二号様式 大量保有報告書", "jplvh020000"],
    [32, "株券等の大量保有の状況の開示に関する内閣府令 第三号様式 大量保有報告書", "jplvh030000"],
    [33, "特定有価証券の内容等の開示に関する内閣府令 臨時報告書", "jpsps000000"],
    [34, "みなし有価証券届出書の追加DEI", "jpsps_みなし有価証券届出書_追加DEI"],
    [35, "特定有価証券の内容等の開示に関する内閣府令 第四号様式 有価証券届出書", "jpsps040000"],
    [36, "特定有価証券の内容等の開示に関する内閣府令 第四号の三様式 有価証券届出書", "jpsps040300"],
    [37, "特定有価証券の内容等の開示に関する内閣府令 第四号の三の二様式 有価証券届出書", "jpsps040302"],
    [38, "特定有価証券の内容等の開示に関する内閣府令 第四号の三の三様式 有価証券届出書", "jpsps040303"],
    [39, "特定有価証券の内容等の開示に関する内閣府令 第五号の二様式 有価証券届出書", "jpsps050200"],
    [40, "特定有価証券の内容等の開示に関する内閣府令 第五号の四様式 有価証券届出書", "jpsps050400"],
    [41, "特定有価証券の内容等の開示に関する内閣府令 第六号様式 有価証券届出書", "jpsps060000"],
    [42, "特定有価証券の内容等の開示に関する内閣府令 第六号の五様式 有価証券届出書", "jpsps060500"],
    [43, "特定有価証券の内容等の開示に関する内閣府令 第六号の七及び第七号様式 有価証券報告書【みなし有価証券届出書】", "jpsps060700"],
    [44, "特定有価証券の内容等の開示に関する内閣府令 第六号の九及び第九号様式 有価証券報告書【みなし有価証券届出書】", "jpsps060900"],
    [45, "特定有価証券の内容等の開示に関する内閣府令 第七号様式 有価証券報告書", "jpsps070000"],
    [46, "特定有価証券の内容等の開示に関する内閣府令 第七号の三様式 有価証券報告書", "jpsps070300"],
    [47, "特定有価証券の内容等の開示に関する内閣府令 第八号の二様式 有価証券報告書", "jpsps080200"],
    [48, "特定有価証券の内容等の開示に関する内閣府令 第八号の四様式 有価証券報告書", "jpsps080400"],
    [49, "特定有価証券の内容等の開示に関する内閣府令 第九号様式 有価証券報告書", "jpsps090000"],
    [50, "特定有価証券の内容等の開示に関する内閣府令 第九号の五様式 有価証券報告書", "jpsps090500"],
    [51, "特定有価証券の内容等の開示に関する内閣府令 第十号様式 半期報告書", "jpsps100000"],
    [52, "特定有価証券の内容等の開示に関する内閣府令 第十号の三様式 半期報告書", "jpsps100300"],
    [53, "特定有価証券の内容等の開示に関する内閣府令 第十一号の二様式 半期報告書", "jpsps110200"],
    [54, "特定有価証券の内容等の開示に関する内閣府令 第十一号の四様式 半期報告書", "jpsps110400"],
    [55, "特定有価証券の内容等の開示に関する内閣府令 第十二号様式 半期報告書", "jpsps120000"],
    [56, "特定有価証券の内容等の開示に関する内閣府令 第十二号の五様式 半期報告書", "jpsps120500"],
    [57, "特定有価証券の内容等の開示に関する内閣府令 第十五号様式 発行登録書", "jpsps150000"],
    [58, "特定有価証券の内容等の開示に関する内閣府令 第十五号の三様式 発行登録書", "jpsps150300"],
    [59, "特定有価証券の内容等の開示に関する内閣府令 第二十一号様式 発行登録追補書類", "jpsps210000"],
    [60, "特定有価証券の内容等の開示に関する内閣府令 第二十五号の三様式 自己株券買付状況報告書", "jpsps250300"],
    [61, "発行者による上場株券等の公開買付けの開示に関する内閣府令 第二号様式 公開買付届出書", "jptoi020000"],
    [62, "発行者による上場株券等の公開買付けの開示に関する内閣府令 第三号様式 公開買付撤回届出書", "jptoi030000"],
    [63, "発行者による上場株券等の公開買付けの開示に関する内閣府令 第四号様式 公開買付報告書", "jptoi040000"],
    [64, "発行者以外の者による株券等の公開買付けの開示に関する内閣府令 第二号様式 公開買付届出書", "jptoo020000"],
    [65, "発行者以外の者による株券等の公開買付けの開示に関する内閣府令 第四号様式 意見表明報告書", "jptoo040000"],
    [66, "発行者以外の者による株券等の公開買付けの開示に関する内閣府令 第五号様式 公開買付撤回届出書", "jptoo050000"],
    [67, "発行者以外の者による株券等の公開買付けの開示に関する内閣府令 第六号様式 公開買付報告書", "jptoo060000"],
    [68, "発行者以外の者による株券等の公開買付けの開示に関する内閣府令 第八号様式 対質問回答報告書", "jptoo08000"]
]

def get_list_by_sheet_name(sheet_name, data):
    for item in data:
        if str(item[0]) == sheet_name:
            return item
    return None

def main():
    # Excelファイルのパス
    file_path = str(sys.argv[1])

    # Excelファイルを読み込む
    workbook = load_workbook(file_path, read_only=True, data_only=True)

    # シートごとに処理を行う関数
    def process_sheet(sheet_name, sheet):
        result_list = get_list_by_sheet_name(sheet_name, data)
        # C列、D列、I列の列番号
        c_col_num = 3  # C列の列番号 (1-indexed)
        d_col_num = 4  # D列の列番号 (1-indexed)
        i_col_num = 9  # I列の列番号 (1-indexed)
        
        # 各列の値を取得
        values = []
        getfuncs=[]
        eliftext=[]
        for i in range(3, sheet.max_row + 1):
            c_value = sheet.cell(row=i, column=c_col_num).value
            i_value = sheet.cell(row=i, column=i_col_num).value
            if c_value !=None:
                values.append(f"\t{i_value} = '' #{c_value}")
                getfuncs.append(f"\tdef get{i_value}(self): #{c_value}\n\t\treturn self.{i_value}\n")
                eliftext.append(f"\t\t\telif fact.concept.qname.localName == '{i_value}': #{c_value}\n\t\t\t\tself.{i_value} = fact.value\n")
    
        # 抽出結果を.pyファイルに書き出す
        output_file_name = f"{result_list[2]}.py"
        classname = str(result_list[2]).split("-")[0]
        with open(output_file_name, 'w', encoding='utf-8') as f:
            f.write(f"from arelle import ModelManager\nfrom arelle import Cntlr\n\n")
            f.write(f"class {classname}:#{result_list[1]}\n")           
            f.write("\n".join(values))
            f.write("\n\n")
            f.write(f"\tdef __init__(self,xbrl_file):\n\t\tctrl = Cntlr.Cntlr()\n\t\tmodel_manager = ModelManager.initialize(ctrl)\n\t\tself.model_xbrl = model_manager.load(xbrl_file)\n")
            f.write(f"\tdef setInfo(self, edinet_info_list):\n")
            f.write(f"\t\tfor fact in self.model_xbrl.facts:\n\t\t\tif fact.concept.qname.localName == 'EDINETCodeDEI':\n\t\t\t\tself.edinet_code = fact.value\n\t\t\t\tfor code_name in edinet_info_list:\n\t\t\t\t\tif code_name[0] == self.edinet_code:\n\t\t\t\t\t\tself.industry_code = code_name[1]\n\t\t\t\t\t\tbreak\n")
            f.write("\n".join(eliftext))
            f.write("\n\n")
            f.write("\n".join(getfuncs))    
        print(f"{sheet_name}の抽出結果を.pyファイルに書き出しました。")

    # シートごとに処理を行う
    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_sheet, sheet_name, workbook[sheet_name]) for sheet_name in workbook.sheetnames]

    # 全ての処理が完了するまで待機
    for future in futures:
        future.result()

    print("すべてのシートの抽出結果を.pyファイルに書き出しました。")
    
    
if __name__ == '__main__':
    main()